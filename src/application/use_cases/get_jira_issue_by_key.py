import base64
import io
import logging

from src.application.ports.jira_port import JiraPort

logger = logging.getLogger(__name__)

# 첨부파일 content 다운로드 크기 제한
_MAX_IMAGE_SIZE = 5 * 1024 * 1024      # 5MB
_MAX_EXCEL_SIZE = 2 * 1024 * 1024      # 2MB
_MAX_TEXT_SIZE = 500 * 1024             # 500KB
_MAX_EXCEL_SHEETS = 3
_MAX_EXCEL_ROWS_PER_SHEET = 200

# MIME 타입 분류
_IMAGE_MIMES = {"image/png", "image/jpeg", "image/gif", "image/webp", "image/svg+xml"}
_EXCEL_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}
_TEXT_MIMES = {
    "text/plain", "text/csv", "text/xml", "text/html",
    "application/json", "application/xml",
}


class GetJiraIssueByKeyUseCase:
    """특정 Jira 이슈를 key(ID)로 조회하는 Use Case"""

    def __init__(self, jira_port: JiraPort):
        self.jira_port = jira_port

    async def execute(self, key: str) -> dict | None:
        """
        Jira 이슈를 key로 조회합니다.

        Args:
            key: Jira 이슈 키 (예: "PROJECT-1234")

        Returns:
            이슈 정보 (dict 형식) 또는 None (이슈가 없을 경우)
        """
        logger.info("🔍 GetJiraIssueByKeyUseCase 실행 시작")
        logger.info("조회할 이슈 키: %s", key)

        # JQL 쿼리 생성
        jql = f'key="{key}"'
        logger.info("생성된 JQL 쿼리: %s", jql)

        issues = await self.jira_port.search_issues(jql)

        if not issues:
            logger.info("이슈를 찾을 수 없음: %s", key)
            return None

        issue = issues[0]
        logger.info("✅ 이슈 조회 완료: %s - %s", issue.key, issue.summary)

        # 첨부파일 조회 및 content 처리
        attachments = await self._process_attachments(issue.key)
        logger.info("첨부파일 처리 완료: %d건", len(attachments))

        return {
            "key": issue.key,
            "summary": issue.summary,
            "status": issue.status,
            "assignee": issue.assignee,
            "description": issue.description,
            "issuetype": issue.issuetype,
            "url": issue.url,
            "custom_fields": dict(issue.custom_fields),
            "attachments": attachments,
        }

    async def _process_attachments(self, issue_key: str) -> list[dict]:
        """첨부파일 메타정보를 조회하고, 지원되는 파일은 content를 다운로드하여 변환한다."""
        try:
            attachments = await self.jira_port.get_issue_attachments(issue_key)
        except Exception as e:
            logger.warning("첨부파일 메타 조회 실패 (%s): %s", issue_key, e)
            return []

        results = []
        for att in attachments:
            entry = {
                "filename": att["filename"],
                "size": att["size"],
                "mimeType": att["mimeType"],
                "content_url": att["content_url"],
                "content_type": "meta_only",  # 기본: 메타정보만
                "content": None,
            }

            mime = att["mimeType"]
            size = att["size"]

            try:
                if mime in _IMAGE_MIMES and size <= _MAX_IMAGE_SIZE:
                    raw = await self.jira_port.download_attachment_content(att["content_url"])
                    entry["content_type"] = "image"
                    entry["content"] = base64.b64encode(raw).decode("ascii")

                elif mime in _EXCEL_MIMES and size <= _MAX_EXCEL_SIZE:
                    raw = await self.jira_port.download_attachment_content(att["content_url"])
                    entry["content_type"] = "text"
                    entry["content"] = self._parse_excel(raw, att["filename"])

                elif (mime in _TEXT_MIMES or att["filename"].endswith((".txt", ".csv", ".log", ".json", ".xml"))) and size <= _MAX_TEXT_SIZE:
                    raw = await self.jira_port.download_attachment_content(att["content_url"])
                    entry["content_type"] = "text"
                    entry["content"] = raw.decode("utf-8", errors="replace")

            except Exception as e:
                logger.warning("첨부파일 content 처리 실패 (%s): %s", att["filename"], e)
                # content 처리 실패 시 메타정보만 유지

            results.append(entry)

        return results

    @staticmethod
    def _parse_excel(raw: bytes, filename: str) -> str:
        """엑셀 바이너리를 텍스트(CSV 형식)로 변환한다."""
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        output_parts = []

        for idx, sheet_name in enumerate(wb.sheetnames):
            if idx >= _MAX_EXCEL_SHEETS:
                output_parts.append(f"\n... 외 {len(wb.sheetnames) - _MAX_EXCEL_SHEETS}개 시트 생략")
                break

            ws = wb[sheet_name]
            output_parts.append(f"=== 시트: {sheet_name} ===")

            row_count = 0
            for row in ws.iter_rows(values_only=True):
                if row_count >= _MAX_EXCEL_ROWS_PER_SHEET:
                    output_parts.append(f"... (행 제한 {_MAX_EXCEL_ROWS_PER_SHEET}행 초과, 나머지 생략)")
                    break
                cells = [str(cell) if cell is not None else "" for cell in row]
                output_parts.append("\t".join(cells))
                row_count += 1

        wb.close()
        return "\n".join(output_parts)
